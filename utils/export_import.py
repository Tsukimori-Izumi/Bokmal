"""CSV/JSON Export and Import utilities."""

import csv
import json
import io
from datetime import date, datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def tasks_to_csv(tasks: list[dict]) -> str:
    """Export tasks to CSV string."""
    output = io.StringIO()
    fieldnames = [
        "id", "wbs", "name", "duration", "start_date", "end_date",
        "progress", "predecessors", "resource_names",
        "is_milestone", "notes"
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for task in tasks:
        row = dict(task)
        # Convert dates
        for key in ("start_date", "end_date"):
            if isinstance(row.get(key), date):
                row[key] = row[key].isoformat()
        writer.writerow(row)
    return output.getvalue()


def csv_to_tasks(csv_text: str) -> list[dict]:
    """Import tasks from CSV string."""
    reader = csv.DictReader(io.StringIO(csv_text))
    tasks = []
    for row in reader:
        task = dict(row)
        # Convert types
        task["id"] = int(task.get("id", 0))
        task["duration"] = int(task.get("duration", 1))
        task["progress"] = float(task.get("progress", 0))
        task["is_milestone"] = task.get("is_milestone", "").lower() == "true"

        for key in ("start_date", "end_date"):
            val = task.get(key, "")
            if val:
                try:
                    task[key] = date.fromisoformat(val)
                except ValueError:
                    task[key] = None
            else:
                task[key] = None
        tasks.append(task)
    return tasks


def project_to_json(project_data: dict) -> str:
    """Export project data to JSON string."""
    def default_serializer(obj):
        if isinstance(obj, date):
            return obj.isoformat()
        raise TypeError(f"Object of type {type(obj)} is not JSON serializable")

    return json.dumps(project_data, default=default_serializer, ensure_ascii=False, indent=2)


def json_to_project(json_text: str) -> dict:
    """Import project data from JSON string."""
    data = json.loads(json_text)

    # Convert date strings back
    if "tasks" in data:
        for task in data["tasks"]:
            for key in ("start_date", "end_date", "constraint_date",
                        "baseline_start", "baseline_end"):
                if task.get(key) and isinstance(task[key], str):
                    try:
                        task[key] = date.fromisoformat(task[key])
                    except ValueError:
                        task[key] = None

    if "project" in data:
        if isinstance(data["project"].get("start_date"), str):
            try:
                data["project"]["start_date"] = date.fromisoformat(data["project"]["start_date"])
            except ValueError:
                pass

    return data


def export_tasks_to_excel(tasks: list[dict], filepath: str) -> bool:
    """Export tasks to an Excel file (.xlsx) using openpyxl.
    Returns True if successful, False if openpyxl is not available.
    """
    if not HAS_OPENPYXL:
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    # Define columns
    columns = [
        "id", "wbs", "name", "duration", "start_date", "end_date",
        "progress", "predecessors", "resource_names",
        "is_milestone", "notes"
    ]

    # Write Header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Write Data
    for row_idx, task in enumerate(tasks, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = task.get(col_name, "")
            # Convert dates to strings or excel date format
            if isinstance(val, date) or isinstance(val, datetime):
                # openpyxl can handle datetimes automatically if desired
                val = val.isoformat()
            elif isinstance(val, bool):
                val = "True" if val else "False"
                
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-adjust column widths roughly
    for col_idx, col_name in enumerate(columns, 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(10, len(col_name) + 2)
        
    # Name column gets a bit more space
    ws.column_dimensions['C'].width = 30
    
    wb.save(filepath)
    return True
